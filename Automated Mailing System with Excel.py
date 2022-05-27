from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import smtplib
from openpyxl import load_workbook

SETTING = {
    'SERVER': 'smtp server address',
    'PORT': port number,
    'USER': 'ID',
    'PASSWORD': 'PW'
}

wb = load_workbook('./example_excel.xlsx', data_only=True) #put excel file directory
ws = wb.active

for row in ws.iter_rows():
    if row[11].value != None:
        addr = row[0].value #put mail address row
        subject = "title"
        content = ("contents")
        attach = row[1].value #put attachment row
        send_email(addr, subject, content, attach)

def email_check(email):
    import re;

    regex = '^[a-zA-Z0-9+-_.]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$'

    if re.match(regex, email):
        return True
    else:
        return False

def send_email(email, subject, content, attach):
        if not email_check(email):
            print('Check mail address => ' + email)
            return

        if attach:
            mime = MIMEMultipart('mixed')
        else:
            mime = MIMEMultipart('alternative')

        mime['From'] = SETTING.get('USER')
        mime['To'] = email
        mime['Subject'] = subject

        contents = content
        text = MIMEText(_text = contents, _charset = 'utf-8')
        mime.attach(text)

        if attach:
            from email.mime.base import MIMEBase
            from email import encoders
            import os

            data = MIMEBase('application', 'octect-stream')
            data.set_payload(open(attach, 'rb').read())
            encoders.encode_base64(data)

            filename = os.path.basename(attach)
            data.add_header('Content-Disposition', 'attachment', filename=('UTF-8', '', filename))
            mime.attach(data)

        smtp = smtplib.SMTP_SSL(SETTING.get('SERVER'), SETTING.get('PORT'))

        smtp.login(SETTING.get('USER'), SETTING.get('PASSWORD'))

        smtp.sendmail(SETTING.get('USER'), email, mime.as_string())

        smtp.close()

        print('Mail Sent Completed ==> ' + email)
